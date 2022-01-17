from bitkub import Bitkub
from configparser import ConfigParser #การเรียกไฟล์ Config
import time
from openpyxl import load_workbook
file='log.xlsx'
lo=load_workbook(file)
log=lo.active
dbconf = ConfigParser()
dbconf.read_file(open(r'config.ini'))
KEY = dbconf.get('Config', 'API_KEY')
PASS = dbconf.get('Config', 'API_SECRET')
Line_Notify = dbconf.get('Config', 'LineNotify')
Asset = dbconf.get('Config', 'Asset').split(",")
coo = dbconf.get('Config', 'Core').split(",")
DCA = dbconf.get('Config', 'DCA')
GAP = dbconf.get('Config', 'GAP')
log['A1']='Asset'
log['A2']='Core'
log['A3']='Befor'
Bot=" "
def cellbe(i):
    return log.cell(8,i+1).value
def cellas(i):
    return log.cell(6,i+1).value
def cellco(i):
    return log.cell(7,i+1).value
try:
    for i in range(len(Asset)):
        if log.cell(1, i + 2).value != Asset[i]:
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])
        elif log.cell(3, i + 2).value != int(coo[i]):
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])

    lo.save(file)
except Exception as e:
    print(e)
    print('Close in 60s')
    time.sleep(60)



asset = log.cell(1,2).value
#Asset =asset.split(",")
core = log.cell(2,2).value
#Core =core.split(",")

Account_name = "JK & L1NEMAN (・ᴥ・) Bot v.2.3.1"
password = ""
from line_notify import LineNotify  # การส่ง Line
notify = LineNotify(Line_Notify)

    # Please note that using RawConfigParser's set functions, you can assign
    # non-string values to keys internally, but will receive an error when
    # attempting to write to a file or when you get it in non-raw mode. Setting
    # values using the mapping protocol or ConfigParser's set() does not allow
    # such assignments to take place.

while True :
    h = 0
    while (h < 24):
        n = 0

        while (n < 60):
            try:
                # เริ่มการทำงานของบอท
                bitkub = Bitkub()
                bitkub.set_api_key(KEY)
                bitkub.set_api_secret(PASS)
                bitkub.status()
                bitkub.servertime()
                res = 'result'
                if Account_name == "":
                    print("\n""Account Name - This is Main Account", ': Broker - ', 'Bitkub')
                else:
                    print("\n"'Account Name - ', Account_name, ': Broker - ', 'Bitkub')

                Get_balance = bitkub.wallet()

                for i in range(len(Asset)):
                    Core = log.cell(2, i + 2).value
                    Asset_01 = Get_balance[res][Asset[i]]
                    AssetName = 'THB_' + Asset[i]
                    get_price = bitkub.ticker(AssetName)
                    Asset_01_Value = Asset_01 * get_price[AssetName]['last']
                    print(Asset_01, Asset[i], '=', "{:.2f}".format(Asset_01_Value), '฿<==>฿', str(Core))

                    rat = get_price[AssetName]['last']
                    CoreAsset = int(Core)
                    GAP = float(GAP)
                    if CoreAsset > 600:
                        if GAP > 2:
                            Rebalance_percent = GAP
                        else:
                            Rebalance_percent = 2

                    else:
                        Rebalance_percent = 1200 / CoreAsset
                    DiffAsset = (CoreAsset * Rebalance_percent / 100)
                    if Asset_01_Value > (CoreAsset + DiffAsset):
                        if DiffAsset > 200:
                            diff_sell = DiffAsset - 4
                        elif DiffAsset > 100:
                            diff_sell = DiffAsset * .98
                        elif DiffAsset > 50:
                            diff_sell = DiffAsset - 1.5
                        else:
                            diff_sell = DiffAsset - 1
                        bitkub.place_ask_by_fiat(sym=AssetName, amt=diff_sell, rat=rat, typ='market')
                        log[str(cellco(i + 1))] = int(Core + 2)  # ขยายพอร์ตเมื่อมีการขาย ทีละ 2฿
                        lo.save(file)  # Save log
                        CoreSell = 'Sell ' + AssetName + ', Core = ฿' + str(Core + 2)
                        notify.send(CoreSell)
                        print("SELL " + str(diff_sell) + " ฿")

                    elif Asset_01_Value < (CoreAsset - (CoreAsset * Rebalance_percent / 100)):
                        diff_buy = CoreAsset * Rebalance_percent / 100
                        bitkub.place_bid(sym=AssetName, amt=diff_buy, rat=rat, typ='market')
                        print("Buy " + str(diff_buy) + " ฿")
                        notify.send('Buy ' + AssetName)

                    else:
                        print('Diff '"{:.2f}".format(Asset_01_Value - CoreAsset), '฿')
                n += 1
                Bot = 'Cash ฿ ' + ('{:.2f}'.format(Get_balance[res]['THB']))
                print(Bot)
                sleep = 60
                time.sleep(sleep)  # Delay for 1 minute (60 seconds).
            except Exception as e:
                print(e)
                try:
                    notify.send(e)
                except Exception as e:
                    print(e)
                    pass
                pass

        h += 1
        notify.send(Bot + "  I am OK.")
    for i in range(len(Asset)):
        Core = log.cell(2, i + 2).value + int(DCA)  # ขยายพอร์ตทุก 24 ชม. (DCA)
        print(Asset[i], str(Core))
        lo.save(file)
        try:
            notify.send(Asset[i] + ' = ฿' + str(Core))
        except Exception as e:
            print(e)
            pass
